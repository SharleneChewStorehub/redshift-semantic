#!/usr/bin/env python3
"""
Comprehensive JSON to Excel Converter for Semantic Layer Documentation
Creates separate Excel files for each JSON with ALL sections included
"""

import json
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import argparse
from datetime import datetime

def style_worksheet(ws, title):
    """Apply professional styling to worksheet"""
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths with reasonable limits
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 15), 150)  # Min 15, Max 150 chars
        ws.column_dimensions[column_letter].width = adjusted_width

def flatten_nested_object(obj, parent_key='', sep='_'):
    """Recursively flatten nested dictionaries and lists"""
    items = []
    
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, (dict, list)):
                items.extend(flatten_nested_object(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
    elif isinstance(obj, list):
        if all(isinstance(item, str) for item in obj):
            # Simple string list
            items.append((parent_key, ', '.join(obj)))
        else:
            # Complex list
            for i, item in enumerate(obj):
                new_key = f"{parent_key}{sep}{i}"
                if isinstance(item, (dict, list)):
                    items.extend(flatten_nested_object(item, new_key, sep=sep).items())
                else:
                    items.append((new_key, item))
    else:
        items.append((parent_key, obj))
    
    return dict(items)

def convert_database_schema(data, output_file):
    """Convert database_schema.json to Excel"""
    print(f"Converting database_schema.json...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_data = []
        
        if 'tables' in data:
            # Create overview sheet
            overview_data = []
            all_columns_data = []
            
            for table_name, columns in data['tables'].items():
                # Overview data
                overview_data.append({
                    'Table Name': table_name,
                    'Column Count': len(columns),
                    'Has Nullable Columns': any(col.get('is_nullable') == 'YES' for col in columns),
                    'Has Default Values': any(col.get('column_default') is not None for col in columns),
                    'Max String Length': max([col.get('character_maximum_length', 0) or 0 for col in columns]),
                    'Data Types': ', '.join(set(col.get('data_type', '') for col in columns))
                })
                
                # All columns data
                for col in columns:
                    col_data = {
                        'Table Name': table_name,
                        'Column Name': col.get('column_name', ''),
                        'Data Type': col.get('data_type', ''),
                        'Is Nullable': col.get('is_nullable', ''),
                        'Column Default': str(col.get('column_default', '')),
                        'Max Length': col.get('character_maximum_length', ''),
                        'Numeric Precision': col.get('numeric_precision', ''),
                        'Numeric Scale': col.get('numeric_scale', ''),
                        'Datetime Precision': col.get('datetime_precision', '')
                    }
                    all_columns_data.append(col_data)
            
            # Create sheets
            overview_df = pd.DataFrame(overview_data)
            overview_df.to_excel(writer, sheet_name='Tables_Overview', index=False)
            ws = writer.sheets['Tables_Overview']
            style_worksheet(ws, 'Tables Overview')
            summary_data.append({'Sheet': 'Tables_Overview', 'Rows': len(overview_df), 'Description': 'Database tables summary'})
            
            all_columns_df = pd.DataFrame(all_columns_data)
            all_columns_df.to_excel(writer, sheet_name='All_Columns', index=False)
            ws = writer.sheets['All_Columns']
            style_worksheet(ws, 'All Columns')
            summary_data.append({'Sheet': 'All_Columns', 'Rows': len(all_columns_df), 'Description': 'All table columns with details'})
            
            # Create individual table sheets for large tables (limit to top 20 by column count)
            top_tables = sorted(data['tables'].items(), key=lambda x: len(x[1]), reverse=True)[:20]
            
            for table_name, columns in top_tables:
                if len(columns) > 10:  # Only create individual sheets for complex tables
                    table_df = pd.DataFrame(columns)
                    safe_name = table_name.replace('/', '_').replace('__', '_')[:31]
                    table_df.to_excel(writer, sheet_name=safe_name, index=False)
                    ws = writer.sheets[safe_name]
                    style_worksheet(ws, table_name)
                    summary_data.append({'Sheet': safe_name, 'Rows': len(table_df), 'Description': f'Columns for {table_name}'})
        
        # Create summary sheet
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            ws = writer.sheets['Summary']
            style_worksheet(ws, 'Summary')
    
    print(f"✓ Created: {output_file} with {len(summary_data)} sheets")
    return len(summary_data)

def convert_data_catalog(data, output_file):
    """Convert data_catalog.json to Excel with ALL sections"""
    print(f"Converting data_catalog.json...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_data = []
        
        # Business Entities - Main overview
        if 'business_entities' in data:
            entities_data = []
            metrics_data = []
            dimensions_data = []
            use_cases_data = []
            
            for entity_name, entity_data in data['business_entities'].items():
                # Main entity info
                entities_data.append({
                    'Entity Name': entity_name,
                    'Description': entity_data.get('description', ''),
                    'Primary Table': entity_data.get('primary_table', ''),
                    'Related Tables': ', '.join(entity_data.get('related_tables', [])),
                    'Use Cases Count': len(entity_data.get('use_cases', [])),
                    'Key Metrics Count': len(entity_data.get('key_metrics', {})),
                    'Has Dimensions': 'dimensions' in entity_data,
                    'Has Event Types': 'event_types' in entity_data
                })
                
                # Metrics detail
                for metric_name, metric_data in entity_data.get('key_metrics', {}).items():
                    metrics_data.append({
                        'Entity': entity_name,
                        'Metric Name': metric_name,
                        'Definition': metric_data.get('definition', ''),
                        'Calculation': metric_data.get('calculation', ''),
                        'Aggregation': metric_data.get('aggregation', ''),
                        'SQL': metric_data.get('sql', ''),
                        'Volume': metric_data.get('volume', ''),
                        'Filter': metric_data.get('filter', ''),
                        'Source Field': metric_data.get('source_field', ''),
                        'Source Table': metric_data.get('source_table', '')
                    })
                
                # Dimensions
                if 'dimensions' in entity_data:
                    for dim_name, dim_data in entity_data['dimensions'].items():
                        dimensions_data.append({
                            'Entity': entity_name,
                            'Dimension': dim_name,
                            'Field': dim_data.get('field', '') if isinstance(dim_data, dict) else '',
                            'Type': dim_data.get('type', '') if isinstance(dim_data, dict) else '',
                            'Values': ', '.join(dim_data.get('values', [])) if isinstance(dim_data, dict) and 'values' in dim_data else str(dim_data),
                            'Description': dim_data.get('description', '') if isinstance(dim_data, dict) else ''
                        })
                
                # Use cases
                for use_case in entity_data.get('use_cases', []):
                    use_cases_data.append({
                        'Entity': entity_name,
                        'Use Case': use_case
                    })
            
            # Create sheets
            for sheet_name, data_list, description in [
                ('Business_Entities', entities_data, 'Business entities overview'),
                ('Key_Metrics', metrics_data, 'All key metrics with calculations'),
                ('Dimensions', dimensions_data, 'Entity dimensions and fields'),
                ('Use_Cases', use_cases_data, 'Use cases by entity')
            ]:
                if data_list:
                    df = pd.DataFrame(data_list)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]
                    style_worksheet(ws, sheet_name)
                    summary_data.append({'Sheet': sheet_name, 'Rows': len(df), 'Description': description})
        
        # Table Definitions
        if 'table_definitions' in data:
            tables_data = []
            fields_data = []
            
            for table_name, table_data in data['table_definitions'].items():
                tables_data.append({
                    'Table Name': table_name,
                    'Description': table_data.get('description', ''),
                    'Record Count': table_data.get('record_count', ''),
                    'High Volume': table_data.get('high_volume', False),
                    'Key Fields Count': len(table_data.get('key_fields', {})),
                    'Has Geographic Data': 'geographic' in table_data.get('description', '').lower(),
                    'Has Financial Data': 'financial' in table_data.get('description', '').lower()
                })
                
                # Key fields detail
                for field_name, field_data in table_data.get('key_fields', {}).items():
                    fields_data.append({
                        'Table': table_name,
                        'Field Name': field_name,
                        'Type': field_data.get('type', ''),
                        'Description': field_data.get('description', ''),
                        'Format': field_data.get('format', ''),
                        'Required': field_data.get('required', ''),
                        'Index': field_data.get('index', ''),
                        'Nullable': field_data.get('nullable', ''),
                        'Examples': ', '.join(field_data.get('examples', [])) if 'examples' in field_data else ''
                    })
            
            for sheet_name, data_list, description in [
                ('Table_Definitions', tables_data, 'Table structure definitions'),
                ('Table_Fields', fields_data, 'Detailed field information')
            ]:
                if data_list:
                    df = pd.DataFrame(data_list)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]
                    style_worksheet(ws, sheet_name)
                    summary_data.append({'Sheet': sheet_name, 'Rows': len(df), 'Description': description})
        
        # Relationship Map
        if 'relationship_map' in data:
            relationships_data = []
            for rel_name, rel_data in data['relationship_map'].items():
                relationships_data.append({
                    'Relationship Name': rel_name,
                    'From': rel_data.get('from', ''),
                    'To': rel_data.get('to', ''),
                    'Type': rel_data.get('type', ''),
                    'Nullable': rel_data.get('nullable', ''),
                    'Description': rel_data.get('description', '')
                })
            
            if relationships_data:
                df = pd.DataFrame(relationships_data)
                df.to_excel(writer, sheet_name='Relationship_Map', index=False)
                ws = writer.sheets['Relationship_Map']
                style_worksheet(ws, 'Relationship Map')
                summary_data.append({'Sheet': 'Relationship_Map', 'Rows': len(df), 'Description': 'Table relationships and joins'})
        
        # Data Quality Patterns
        if 'data_quality_patterns' in data:
            quality_data = []
            for pattern_name, pattern_data in data['data_quality_patterns'].items():
                flattened = flatten_nested_object(pattern_data)
                for key, value in flattened.items():
                    quality_data.append({
                        'Pattern Category': pattern_name,
                        'Property': key,
                        'Value': str(value)
                    })
            
            if quality_data:
                df = pd.DataFrame(quality_data)
                df.to_excel(writer, sheet_name='Data_Quality_Patterns', index=False)
                ws = writer.sheets['Data_Quality_Patterns']
                style_worksheet(ws, 'Data Quality Patterns')
                summary_data.append({'Sheet': 'Data_Quality_Patterns', 'Rows': len(df), 'Description': 'Data quality rules and patterns'})
        
        # Create summary sheet
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            ws = writer.sheets['Summary']
            style_worksheet(ws, 'Summary')
    
    print(f"✓ Created: {output_file} with {len(summary_data)} sheets")
    return len(summary_data)

def convert_business_rules(data, output_file):
    """Convert business_rules.json to Excel with ALL sections"""
    print(f"Converting business_rules.json...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_data = []
        
        # Process each main section
        sections_to_process = [
            'mandatory_filters', 'data_quality_rules', 'performance_constraints',
            'access_control', 'calculation_standards', 'error_handling',
            'default_behaviors', 'validation_rules', 'marketplace_rules',
            'payment_validation_rules', 'tax_validation_rules', 'employee_access_rules',
            'promotion_rules', 'loyalty_program_rules', 'payment_gateway_rules',
            'shipping_rules', 'inventory_rules', 'feedback_rules',
            'campaign_rules', 'regional_rules', 'table_specific_constraints'
        ]
        
        for section_name in sections_to_process:
            if section_name in data:
                section_data = []
                
                def process_rules(rules_dict, category='', subcategory=''):
                    for key, value in rules_dict.items():
                        if isinstance(value, dict):
                            if any(k in value for k in ['rule', 'filter', 'requirement', 'validation']):
                                # This is a rule definition
                                section_data.append({
                                    'Category': category,
                                    'Subcategory': subcategory,
                                    'Rule Name': key,
                                    'Rule/Filter': value.get('rule', value.get('filter', value.get('requirement', value.get('validation', '')))),
                                    'Description': value.get('description', ''),
                                    'Applies To': ', '.join(value.get('applies_to', [])) if isinstance(value.get('applies_to'), list) else str(value.get('applies_to', '')),
                                    'Error Message': value.get('error_if_missing', value.get('error_message', '')),
                                    'Reason': value.get('reason', ''),
                                    'Note': value.get('note', value.get('implementation_note', '')),
                                    'Severity': value.get('severity', ''),
                                    'Examples': ', '.join(value.get('examples', [])) if 'examples' in value else '',
                                    'Default': value.get('default', ''),
                                    'Override For': ', '.join(value.get('override_for', [])) if 'override_for' in value else ''
                                })
                            else:
                                # This is a category, recurse
                                process_rules(value, category or key, key if category else '')
                        else:
                            # Simple key-value pair
                            section_data.append({
                                'Category': category,
                                'Subcategory': subcategory,
                                'Rule Name': key,
                                'Rule/Filter': '',
                                'Description': str(value),
                                'Applies To': '',
                                'Error Message': '',
                                'Reason': '',
                                'Note': '',
                                'Severity': '',
                                'Examples': '',
                                'Default': '',
                                'Override For': ''
                            })
                
                process_rules(data[section_name])
                
                if section_data:
                    df = pd.DataFrame(section_data)
                    sheet_name = section_name.title().replace('_', '_')[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]
                    style_worksheet(ws, sheet_name)
                    summary_data.append({'Sheet': sheet_name, 'Rows': len(df), 'Description': f'{section_name.replace("_", " ").title()} rules'})
        
        # Create summary sheet
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            ws = writer.sheets['Summary']
            style_worksheet(ws, 'Summary')
    
    print(f"✓ Created: {output_file} with {len(summary_data)} sheets")
    return len(summary_data)

def convert_value_mappings(data, output_file):
    """Convert value_mappings.json to Excel with ALL mappings"""
    print(f"Converting value_mappings.json...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_data = []
        
        # Overview of all mapping categories
        overview_data = []
        all_mappings_data = []
        
        for field_name, field_data in data.items():
            if field_name in ['version', 'description', 'last_updated', 'created_from']:
                continue
            
            if isinstance(field_data, dict):
                mappings_count = len(field_data.get('mappings', {}))
                overview_data.append({
                    'Field Name': field_name,
                    'Field Type': field_data.get('type', ''),
                    'Mappings Count': mappings_count,
                    'Has Discovered Values': 'discovered_values' in field_data,
                    'Has Type Info': 'type' in field_data,
                    'Description': field_data.get('description', '')
                })
                
                # Detailed mappings
                if 'mappings' in field_data:
                    for code, mapping in field_data['mappings'].items():
                        if isinstance(mapping, dict):
                            flattened = flatten_nested_object(mapping)
                            row = {
                                'Field': field_name,
                                'Code': code,
                                'Display Name': mapping.get('display_name', ''),
                                'Category': mapping.get('category', ''),
                                'Is Active': mapping.get('is_active', ''),
                                'Description': mapping.get('description', ''),
                                'Currency': mapping.get('currency', ''),
                                'Country': mapping.get('country', ''),
                                'Timezone': mapping.get('timezone', ''),
                                'Language': mapping.get('language', ''),
                                'Symbol': mapping.get('symbol', ''),
                                'Normalized To': mapping.get('normalized_to', ''),
                                'All Properties': str(flattened)
                            }
                        else:
                            row = {
                                'Field': field_name,
                                'Code': code,
                                'Display Name': str(mapping),
                                'Category': '',
                                'Is Active': '',
                                'Description': '',
                                'Currency': '',
                                'Country': '',
                                'Timezone': '',
                                'Language': '',
                                'Symbol': '',
                                'Normalized To': '',
                                'All Properties': str(mapping)
                            }
                        all_mappings_data.append(row)
        
        # Create sheets
        for sheet_name, data_list, description in [
            ('Mappings_Overview', overview_data, 'Overview of all mapping categories'),
            ('All_Mappings', all_mappings_data, 'Detailed mappings for all fields')
        ]:
            if data_list:
                df = pd.DataFrame(data_list)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                style_worksheet(ws, sheet_name)
                summary_data.append({'Sheet': sheet_name, 'Rows': len(df), 'Description': description})
        
        # Create individual sheets for complex mappings
        complex_fields = ['country', 'currency', 'channel', 'order_status', 'loyalty_event_type', 'tax_code']
        
        for field_name in complex_fields:
            if field_name in data and 'mappings' in data[field_name]:
                field_data = []
                for code, mapping in data[field_name]['mappings'].items():
                    if isinstance(mapping, dict):
                        row = {'Code': code}
                        row.update(flatten_nested_object(mapping))
                    else:
                        row = {'Code': code, 'Value': str(mapping)}
                    field_data.append(row)
                
                if field_data:
                    df = pd.DataFrame(field_data)
                    sheet_name = field_name.title()[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]
                    style_worksheet(ws, sheet_name)
                    summary_data.append({'Sheet': sheet_name, 'Rows': len(df), 'Description': f'{field_name} detailed mappings'})
        
        # Create summary sheet
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            ws = writer.sheets['Summary']
            style_worksheet(ws, 'Summary')
    
    print(f"✓ Created: {output_file} with {len(summary_data)} sheets")
    return len(summary_data)

def convert_query_patterns(data, output_file):
    """Convert query_patterns.json to Excel with ALL pattern categories"""
    print(f"Converting query_patterns.json...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_data = []
        
        # Process all pattern categories
        pattern_categories = [
            'core_patterns', 'advanced_patterns', 'optimization_patterns',
            'composition_rules', 'multi_channel_patterns', 'tax_patterns',
            'employee_patterns', 'promotion_patterns', 'order_lifecycle_patterns',
            'inventory_patterns', 'financial_reconciliation_patterns',
            'voucher_patterns', 'marketing_campaign_patterns'
        ]
        
        for category in pattern_categories:
            if category in data:
                patterns_data = []
                
                for pattern_name, pattern_data in data[category].items():
                    if isinstance(pattern_data, dict):
                        row = {
                            'Pattern Name': pattern_name,
                            'Purpose': pattern_data.get('purpose', ''),
                            'Template': pattern_data.get('template', ''),
                            'Logic': pattern_data.get('logic', ''),
                            'Use Cases': ', '.join(pattern_data.get('use_cases', [])),
                            'Performance Notes': pattern_data.get('performance_notes', ''),
                            'Optimization': pattern_data.get('optimization', ''),
                            'Insights': pattern_data.get('insights', ''),
                            'Requirements': pattern_data.get('requirements', ''),
                            'Description': pattern_data.get('description', ''),
                            'Example': pattern_data.get('example', ''),
                            'Joins': ', '.join(pattern_data.get('joins', [])),
                            'Metrics': ', '.join(pattern_data.get('metrics', [])),
                            'KPIs': ', '.join(pattern_data.get('kpis', [])),
                            'Types': ', '.join(pattern_data.get('types', [])),
                            'Channels': ', '.join(pattern_data.get('channels', [])),
                            'Gateways': ', '.join(pattern_data.get('gateways', [])),
                            'Tracks': ', '.join(pattern_data.get('tracks', [])),
                            'Alerts': ', '.join(pattern_data.get('alerts', []))
                        }
                        
                        # Add parameters info
                        if 'parameters' in pattern_data:
                            params = []
                            for param, desc in pattern_data['parameters'].items():
                                params.append(f"{param}: {desc}")
                            row['Parameters'] = '; '.join(params)
                        else:
                            row['Parameters'] = ''
                        
                        # Add rules if present
                        if 'rules' in pattern_data:
                            row['Rules'] = ', '.join(pattern_data['rules'])
                        else:
                            row['Rules'] = ''
                        
                        patterns_data.append(row)
                    else:
                        # Handle simple string values
                        patterns_data.append({
                            'Pattern Name': pattern_name,
                            'Purpose': '',
                            'Template': '',
                            'Logic': str(pattern_data),
                            'Use Cases': '',
                            'Performance Notes': '',
                            'Optimization': '',
                            'Insights': '',
                            'Requirements': '',
                            'Description': '',
                            'Example': '',
                            'Joins': '',
                            'Metrics': '',
                            'KPIs': '',
                            'Types': '',
                            'Channels': '',
                            'Gateways': '',
                            'Tracks': '',
                            'Alerts': '',
                            'Parameters': '',
                            'Rules': ''
                        })
                
                if patterns_data:
                    df = pd.DataFrame(patterns_data)
                    sheet_name = category.replace('_patterns', '').replace('_', '_').title()[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]
                    style_worksheet(ws, sheet_name)
                    summary_data.append({'Sheet': sheet_name, 'Rows': len(df), 'Description': f'{category.replace("_", " ").title()} patterns'})
        
        # Create summary sheet
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            ws = writer.sheets['Summary']
            style_worksheet(ws, 'Summary')
    
    print(f"✓ Created: {output_file} with {len(summary_data)} sheets")
    return len(summary_data)

def convert_semantic_context(data, output_file):
    """Convert semantic_context.json to Excel with ALL sections"""
    print(f"Converting semantic_context.json...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_data = []
        
        # Process all main sections
        sections = [
            'default_interpretations', 'disambiguation_rules', 'regional_context',
            'intent_patterns', 'confidence_scoring', 'marketplace_integration',
            'data_patterns', 'multi_channel_insights', 'operational_insights',
            'learning_patterns'
        ]
        
        for section_name in sections:
            if section_name in data:
                section_data = []
                
                def process_context_section(ctx_dict, category='', subcategory=''):
                    for key, value in ctx_dict.items():
                        if isinstance(value, dict):
                            if any(k in value for k in ['default', 'mapping', 'rule', 'pattern']):
                                # This is a definition
                                section_data.append({
                                    'Category': category,
                                    'Subcategory': subcategory,
                                    'Item': key,
                                    'Default': value.get('default', ''),
                                    'Reason': value.get('reason', ''),
                                    'Description': value.get('description', ''),
                                    'Alternatives': ', '.join(value.get('alternatives', [])) if 'alternatives' in value else '',
                                    'Values': ', '.join(value.get('values', [])) if 'values' in value else '',
                                    'Type': value.get('type', ''),
                                    'Symbol': value.get('symbol', ''),
                                    'Code': value.get('code', ''),
                                    'Pattern': value.get('pattern', ''),
                                    'Rule': value.get('rule', ''),
                                    'Weight': value.get('weight', ''),
                                    'Confidence': value.get('confidence', ''),
                                    'All Properties': str(value)
                                })
                            else:
                                # This is a subcategory, recurse
                                process_context_section(value, category or key, key if category else '')
                        elif isinstance(value, list):
                            section_data.append({